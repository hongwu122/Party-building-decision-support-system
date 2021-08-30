# import traceback
# from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import *
from tkinter import ttk
from tkinter.ttk import *
from tkinter import scrolledtext

from faker import Faker
import webbrowser
import os
import random
import time
import datetime
import openpyxl
from openpyxl.styles import Color, Font, Alignment, Border, Side, PatternFill, colors
from openpyxl.utils import get_column_letter
# 通用————合并表格（暂时只针对xlsx文件格式）
def general_merge_book(path, filename, biaoti_row=0, biaotou_row=1, lishi_row=False):
    # 默认大标题行0行，即没有大标题行。  biaoti_row==1，即第一行是大标题，  biaoti_row==2，即前两行是大标题。
    # 默认表头行在在第一行。 biaotou_row==0行，即没有表头， biaotou_row=2在第二行
    # 没有例示行，如果例示行==True，那就例示行默认在表头行下面一行
    if path == "":
        messagebox.showinfo("提示","请输入正确的需合并文件的文件夹路径！")
        scr_output(scr_0,'\n本次没有正确输入正确的需合并文件的文件夹路径！ \n请把合并文件全部放在一个文件夹里面，并选择该文件夹\n\n')

    if path != "":
        try:
            # # print(os.listdir('./test')) # 打印test目录下所有文件
            print('当前工作路径',os.path.abspath('.')) # 打印当前目录
            # # xlsx_files = [x for x in os.listdir('./test') if os.path.isfile(x) and os.path.splitext(x)[1] == '.xlsx'] # 罗列当前目录内所有xlsx文件
            xlsx_files = [x for x in os.listdir(path) if os.path.splitext(x)[1] == '.xlsx']  # 罗列当前目录内所有xlsx文件
            scr_output(scr_0,'\n\n需要统计{}个表格'.format(len(xlsx_files)))
            scr_output(scr_0,'\n\n统计表格有：\n{}'.format(xlsx_files))
            print('需要统计', len(xlsx_files), '个表格')
            print('统计表格有：\n', xlsx_files)  # 本目录下的xlsx文件名字列表
            print(xlsx_files[0])

            xlsx_file_0 = path + '/' + xlsx_files[0]
            data_biaotou = []  # 用来复制表头数据
            workbook_0 = openpyxl.load_workbook(filename=xlsx_file_0)
            worksheet_0 = workbook_0.active
            if biaotou_row != 0: # 表头存在，才复制
                # 复制表头数据
                for col in range(1, worksheet_0.max_column + 1):
                    data_biaotou.append(worksheet_0.cell(row=biaotou_row, column=col).value) # 默认表头在第一行
                # data_1 = worksheet_0.cell(row=1, column=1).value
                # data_2 = worksheet_0.cell(row=2, column=1).value
                print('表头', data_biaotou)
                scr_output(scr_0,'\n\n存在表头：\n{}'.format(data_biaotou))

            data_shuju = []  # 用来复制储存数据的数据集
            num = len(xlsx_files)
            print('\n\n开始提取数据…………\n')

            # 判断储存行开始位置
            if biaotou_row == 0:  # 没有表头，储存行开始在标题行之后
                cucun_row = biaoti_row + 1
            else:  # 有表头行，要多加一行才开始储存
                cucun_row = biaoti_row + 2
            # 遍历每个文件
            for n in range(num):
                xlsx_file = path + '/' + xlsx_files[n]
                workbook_n = openpyxl.load_workbook(filename=xlsx_file)
                worksheet_n = workbook_n.active

                if lishi_row: # 存在合并文件中存在例示行，需要检测出来，删掉，不要储存进来我们的数据集
                    # 删除空行（涉及删除行后，行数的索引值发生变化，所以复杂了点）
                    # 前三行如果是None，就判定为空行，或者假设第二列名字出现了张三这个人
                    zero_row_list = []
                    for row in range(1, worksheet_n.max_row + 1):
                        if (worksheet_n.cell(row, 1).value == 'None' and worksheet_n.cell(row,2).value == 'None' and worksheet_n.cell(row, 3).value == 'None') or \
                            (worksheet_n.cell(row, 1).value == None and worksheet_n.cell(row,2).value == None and worksheet_n.cell(row, 3).value == None) or \
                            worksheet_n.cell(row, 2).value == '张三':
                            zero_row_list.append(row)
                    delete_row = 0
                    # print('\n\n开始删除{}的空行…………\n'.format(xlsx_files[n]))
                    scr_output(scr_0,'\n\n开始删除{}的空行/空列和张三示例行…………\n'.format(xlsx_files[n]))
                    for j in zero_row_list:
                        # scr_output(scr_0,'\n原本该空值在第{}行,正在删除该空值目前所在的第{}行'.format(j , j -delete_row)
                        # print('原本该空值在第{}行,正在删除该空值目前所在的第{}行'.format(j , j -delete_row))
                        worksheet_n.delete_rows(j - delete_row, 1)
                        delete_row = delete_row + 1

                # scr_output(scr_0,'\n\n正在删除空列（第十七列后的五列数）\n')
                # 删除空列，第十七列后的列数
                # worksheet_n.delete_cols(17, 5)

                # 储存数据
                for i in range(cucun_row, worksheet_n.max_row + 1):
                    list = []
                    for j in range(1, worksheet_n.max_column + 1):
                        list.append(worksheet_n.cell(row=i, column=j).value)
                    data_shuju.append(list)

            # print('数据', data_shuju)
            time.sleep(1)
            data2 = '\n\n数据：\n'
            for i in data_shuju: # 用来输出窗口显示数据更好看
                data2 = data2 + str(i) + '\n'
            scr_output(scr_0,data2)
            # 自此，以上代码获取了全部文件的有效数据

            # # 汇总表头和数据,新建保存总表

            # 不需要添加表头，有模板表了，以下代码注释
            # data = []
            # data.append(data_biaotou)  # 添加表头
            # for l in range(len(data_shuju)):  # 添加数据
            #     data.append(data_shuju[l])
            workbook = openpyxl.load_workbook(xlsx_file_0)  # 打开模板表，以第一个文件作为模板
            worksheet = workbook.worksheets[0]
            worksheet.title = '汇总'
            try:
                workbook.remove(workbook.worksheets[1:])
            except:
                print("后面的工作表移除失败")
            # 写入大标题
            # worksheet.cell(1,1).value = worksheet_0.cell(row=1, column=1).value
            # 写入第二行表头
            # worksheet.cell(biaotou_row, 1).value = worksheet_0.cell(row=biaotou_row, column=1).value
            # worksheet.merge_cells(start_column=1, end_column=15, start_row=1, end_row=1)  # 合并单元格
            # worksheet.merge_cells(start_column=1, end_column=15, start_row=2, end_row=2)

            # 删除这个表的除了大标题行和表头行的全部没用数据，即储存行开始的数据
            for row in range(0, worksheet.max_row): # 懒得计算，索性多删几行
                worksheet.delete_rows(cucun_row+row, 1)

            # 写入数据
            for n_row in range(0, len(data_shuju)):  # 从第三行开始写入数据
                for n_col in range(0, len(data_shuju[n_row])):
                    worksheet.cell(row=n_row + cucun_row, column=n_col+1, value=str(data_shuju[n_row][n_col]))

            # 获取四个区域
            max_row = worksheet.max_row  # 获得最大行数
            max_column = worksheet.max_column  # 获得最大列数
            min_row = worksheet.min_row
            min_column = worksheet.min_column

            scr_output(scr_0,'\n\n数据写入总表完成…………\n')
            scr_output(scr_0,'\n\n写入期数表头总表完成…………\n')
            scr_output(scr_0,'\n\n开始给区域设置设置框线…………\n')
            scr_output(scr_0,'\n\n开始居中对齐…………\n')
            # 给区域设置设置框线
            for row in tuple(worksheet[min_row+2:max_row]):
                for cell in row:
                    cell.border = my_border('thin', 'thin', 'thin', 'thin')
                    # 设置单元格对齐方式 Alignment(horizontal=水平对齐模式,vertical=垂直对齐模式,text_rotation=旋转角度,wrap_text=是否自动换行)
                    cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)

            # 区域自动调整列宽
            column_widths = []  # 定义用来获取当前列最大宽度的空列表
            for i, col in enumerate(
                    worksheet.iter_cols(min_col=min_column, max_col=max_column, min_row=min_row, max_row=max_row)):
                for cell in col:
                    value = cell.value
                    if value is not None:
                        if isinstance(value, str) is False:
                            value = str(value)
                        try:
                            column_widths[i] = max(column_widths[i], len(value))
                        except IndexError:
                            column_widths.append(len(value))
            # print('column_widths', column_widths)  # 得到该列最大的一个单元格的宽度（字符串数量）
            for i, width in enumerate(column_widths):
                col_name = get_column_letter(min_column + i)  # 获取行字母表头
                value = column_widths[i] * 2  # 设置列宽为最大长度比例
                worksheet.column_dimensions[col_name].width = value

            workbook.save(filename=filename + '.xlsx')  # 保存xlsx
            print('文件夹内全部文件合并完成')
            messagebox.showinfo('小提示', '文件夹内全部文件合并完成 成功！')
            scr_output(scr_0,'\n\n文件夹内全部文件合并完成 成功！\n')
            scr_output(scr_0, '\n保存的文件路径为：\n{}\n\n\n\n\n\n'.format(
                '/'.join(path.split('/')[:-1]) + '/' + pathin6_0.get() + '.xlsx'))
            # AttributeError: 'MergedCell' object attribute 'value' is read-only
            # 读取到了合并的单元格，报错

        except Exception as error:
            error = str(error)
            print('错误提示', error)
            scr_output(scr_0, '\n合并文件 失败！\n\n\n本次错误信息：\n{}'.format(error))
            scr_output(scr_0, '\n--------文件没有成功保存--------\n\n\n\n\n\n\n')
            messagebox.showinfo('错误提示', '合并文件 失败！\n错误信息：\n{}'.format(error))


