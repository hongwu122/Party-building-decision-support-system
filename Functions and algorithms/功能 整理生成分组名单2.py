from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


def auto_column(workbook_name):

    wb = openpyxl.load_workbook(workbook_name)
    ws = wb.active

    # 第一步：计算每列最大宽度，并存储在列表lks中。

    lks = []  # 英文变量太费劲，用汉语首字拼音代替
    for i in range(1, ws.max_column + 1):  # 每列循环
        lk = 1  # 定义初始列宽，并在每个行循环完成后重置
        for j in range(1, ws.max_row + 1):  # 每行循环
            sz = ws.cell(row=j, column=i).value  # 每个单元格内容
            if isinstance(sz, str):  # 中文占用多个字节，需要分开处理
                lk1 = len(sz.encode('gbk'))  # gbk解码一个中文两字节，utf-8一个中文三字节，gbk合适
            else:
                lk1 = len(str(sz))
            if lk < lk1:
                lk = lk1  # 借助每行循环将最大值存入lk中
            # print(lk)
        lks.append(lk)  # 将每列最大宽度加入列表。（犯了一个错，用lks = lks.append(lk)报错，append会修改列表变量，返回值none，而none不能继续用append方法）

    # 第二步：设置列宽
    for i in range(1, ws.max_column + 1):
        k = get_column_letter(i)  # 将数字转化为列名,26个字母以内也可以用[chr(i).upper() for i in range(97, 123)]，不用导入模块
        ws.column_dimensions[k].width = lks[i - 1] + 2  # 设置列宽，一般加两个字节宽度，可以根据实际情况灵活调整

    wb.close()
    wb.save(workbook_name)

def fenzu_main2(workbook_name,output_workbook_name):
    # 加载原始Excel文件
    workbook = openpyxl.load_workbook(workbook_name)
    # 创建新的Excel文件
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    # new_sheet.title = '分组名单'
    group_num = 1

    # 初始化开始写入的行数和列数
    current_row = 3
    current_column = 1
    cur_row = 3
    down_row1 = 3
    down_row2 = 3
    # 复制每个sheet的内容到新的sheet中
    for i, sheet in enumerate(workbook.worksheets):
        group_num = i
        print(i,sheet) # 0 <Worksheet "第一组">
        # 判断sheet是奇数个还是偶数个
        if (i+1) % 2 == 0:
            current_column = 7
            current_row = cur_row

        else:
            current_column = 1

        cur_row = current_row
        # 从第三行开始复制
        ori_row = 3
        ori_col = 1
        if sheet.cell(8,3).value ==None:
            print("该sheet为空")
            continue
        for row in sheet.iter_rows(min_row=3, values_only=True):
            # 忽略前两列
            for value in row:
                if value != None:
                    value = value.replace(" ",'')
                new_sheet.cell(row=current_row, column=current_column, value=value)
                original_cell = sheet.cell(row=ori_row, column=ori_col)

                new_cell = new_sheet.cell(row=current_row, column=current_column, value=value)
                # 设置单元格居中对齐，宋体字体
                new_cell.font = copy(original_cell.font)
                new_cell.alignment = copy(original_cell.alignment)
                new_cell.border = copy(original_cell.border)
                new_cell.number_format = copy(original_cell.number_format)
                new_cell.fill = copy(original_cell.fill)

                # 合并单元格
                if original_cell.coordinate in sheet.merged_cells:
                    new_sheet.merge_cells(new_cell.coordinate)

                current_column += 1
                ori_col = ori_col + 1

            # 每行写完后回到第一列/第七列
            if (i + 1) % 2 == 0:
                current_column = 7
            else:
                current_column = 1
            current_row += 1
            ori_row = ori_row + 1
            ori_col = 1

        if (i+1) % 2 != 0:
            down_row1 = current_row
        else:
            down_row2 = current_row
        # 每个sheet复制完后，将行数加上2
        current_row = max(down_row1,down_row2) + 2

    # 保存新的Excel文件
    new_workbook.save(output_workbook_name)
    return group_num

def fenzu_biaotou(workbook_name,new_workbook_name,group_num):
    # 加载原始Excel文件
    new_workbook = openpyxl.load_workbook(new_workbook_name)
    workbook = openpyxl.load_workbook(workbook_name)
    new_sheet = new_workbook.active
    # 写入表头
    new_sheet.cell(1, 1, workbook.worksheets[0].cell(1, 1).value).alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)
    new_sheet.cell(2, 1, workbook.worksheets[0].cell(2, 1).value).alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)
    if group_num >= 2:
        end_column = 11
    else:
        end_column = 5
    new_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    new_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)

    # 保存新的Excel文件
    new_workbook.save(new_workbook_name)

group_num = fenzu_main2('original.xlsx','merged.xlsx')
auto_column(workbook_name='merged.xlsx')
fenzu_biaotou('original.xlsx','merged.xlsx',group_num)

# # 忽略前两列
# for value in row:
#     if value != None:
#         value = value.replace(" ", '')
#     new_sheet.cell(row=current_row, column=current_column, value=value)
#     # new_sheet.column_dimensions[get_column_letter(j + 1)].width = sheet.column_dimensions[get_column_letter(j + 1)].width
#
#     original_cell = sheet.cell(row=ori_row, column=ori_col)
#
#     new_cell = new_sheet.cell(row=current_row, column=current_column, value=value)
#     # 设置单元格居中对齐，宋体字体
#     new_cell.fill = copy(original_cell.fill)
#     # if original_cell.has_style:
#     new_cell.font = copy(original_cell.font)
#     new_cell.alignment = copy(original_cell.alignment)
#     new_cell.border = copy(original_cell.border)
#     # new_cell.number_format = copy(original_cell.number_format)
#     # new_cell._style = copy(original_cell._style)
#     # new_cell.protection = copy(original_cell.protection)
#
#     # # 合并单元格
#     # if original_cell.coordinate in sheet.merged_cells:
#     #     new_sheet.merge_cells(original_cell.coordinate)