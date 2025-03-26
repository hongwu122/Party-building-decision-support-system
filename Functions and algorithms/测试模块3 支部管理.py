import openpyxl


filepath = "mould//模板0 学员花名册.xlsx"

if os.path.splitext(filepath)[1] == '.xls':  # 说明是xls文件
    xls_to_xlsx(path=filepath, sole=True)  # 给路径，让其自己转换成xlsx的
    scr_output(scr3_11, '\n\n检测到本文件是xls格式文件，已经自动在原路径转换生成可读取的xlsx文件类型：\n')
    path = os.path.splitext(filepath)[0] + '.xlsx'
workbook = openpyxl.load_workbook(filepath)
new_workbook = openpyxl.Workbook()
worksheet = workbook.worksheets[0]
new_worksheet = new_workbook.worksheets[0]

a,b,c,d,e,f = 0,0,0,0,0,0
biaotou_r = ['3','2','1']
for br in biaotou_r:
    bt_row = int(br)
    one_cell = worksheet[br]  # 获取第1行的数据
    for i in one_cell:
        # print(i.value)
        # print(i.column)
        if '姓名' in i.value:
            a = i.column
        if '出生' in i.value:
            b = i.column
        if '申请书' in i.value:
            c = i.column
        if  ('专业' or "班级") in i.value:
            d = i.column
        if '支部' in i.value:
            e = i.column
        if '备注' in i.value:
            f = i.column
    if a != 0 and b != 0 and c != 0 and d != 0 and e != 0 and f != 0:
        break
    else: # if a == 0 or b == 0 or c == 0 or d == 0 or e == 0 or f == 0:
        if br == '1':
            print("没有找到全部表头，请检查导入的名单文件！")
            # messagebox.showinfo("提示", "没有找到全部表头，请检查导入的名单文件！")
            # scr_output(scr3_11, '没有找到全部表头，请检查导入的名单文件！')
print(a,b,c,d,e,f)
# 写入表头
biaotou = ['序号','姓名','出生日期','申请书日期','推荐两名党员','推荐群团组织','推荐党支部/党小组','备注']
for bt in range(len(biaotou)):  # 添加数据
    new_worksheet.cell(row=1, column=bt+1, value=str(biaotou[bt]))

r_count = 2
for r in range(worksheet.max_row):
    if worksheet.cell(bt_row+1,a).value == None:
        continue
    new_worksheet.cell(row=r_count, column=1, value=str(r_count-1))
    # 写入姓名
    new_worksheet.cell(row=r_count, column=2, value=str(worksheet.cell(bt_row+1,a).value))
    # 写入出生日期
    print(b)
    print(worksheet.cell(bt_row+1,b).value)
    new_worksheet.cell(row=r_count, column=3, value=str(worksheet.cell(bt_row+1,b).value))
    # 写入申请书日期
    new_worksheet.cell(row=r_count, column=4, value=str(worksheet.cell(bt_row+1,c).value))
    # 写入推荐两名党员
    new_worksheet.cell(row=r_count, column=5, value=str('(请注意自行填写)'))
    # 写入推荐群团组织
    new_worksheet.cell(row=r_count, column=6, value=str(worksheet.cell(bt_row+1,d).value)+'团支部')
    # 写入推荐党支部/党小组
    new_worksheet.cell(row=r_count, column=7, value=str(worksheet.cell(bt_row+1,e).value)) # 全称
    # 写入备注
    new_worksheet.cell(row=r_count, column=8, value=str(worksheet.cell(bt_row+1,f).value))


scr_output(scr3_11,'\n\n开始给区域设置设置框线…………\n')
scr_output(scr3_11,'\n\n开始居中对齐…………\n')
# 给区域设置设置框线
for row in tuple(worksheet[new_worksheet.min_row:new_worksheet.max_row]):
    for cell in row:
        cell.border = my_border('thin', 'thin', 'thin', 'thin')
        # 设置单元格对齐方式 Alignment(horizontal=水平对齐模式,vertical=垂直对齐模式,text_rotation=旋转角度,wrap_text=是否自动换行)
        alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)
        cell.alignment = alignment

scr_output(scr3_11,'\n\n开始区域自动调整列宽…………\n')
# 区域自动调整列宽]

column_widths = []  # 定义用来获取当前列最大宽度的空列表
for i, col in enumerate(worksheet.iter_cols(min_col=new_worksheet.min_column, max_col=new_worksheet.max_column, min_row=new_worksheet.min_row)):
    for cell in col:
        value = cell.value
        if value is not None:
            if isinstance(value, str) is False:
                value = str(value)
            try:
                column_widths[i] = max(column_widths[i], len(value))
            except IndexError:
                column_widths.append(len(value))
# print('column_widths', column_widths)  # 得到该列最大的一个单元格的宽度（字符串数量）
for i, width in enumerate(column_widths):
    col_name = get_column_letter(new_worksheet.min_column + i)  # 获取行字母表头
    value = column_widths[i] * 2 # 设置列宽为最大长度比例
    worksheet.column_dimensions[col_name].width = value
                
new_workbook.save("积极分子备案报告名单.xlsx")

