# 定义边框样式
def my_border(t_border, b_border, l_border, r_border):
    border = Border(top=Side(border_style=t_border, color=colors.BLACK),
                    bottom=Side(border_style=b_border, color=colors.BLACK),
                    left=Side(border_style=l_border, color=colors.BLACK),
                    right=Side(border_style=r_border, color=colors.BLACK))
    return border

# import pandas as pd
import openpyxl
from openpyxl.styles import Color, Font, Alignment, Border, Side, PatternFill, colors
# import numpy as np

def fenzu_smart(path,save):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.worksheets[0]
    # 定义小组空列表
    data = []
    # 遍历本小组每个人信息，收集成列表
    minrow = worksheet.min_row  # 最小行
    maxrow = worksheet.max_row  # 最大行
    mincol = worksheet.min_column  # 最小列
    maxcol = worksheet.max_column  # 最大列

    # 遍历获取相关列是否存在
    for row in tuple(worksheet[1:5]):  # 在前1到5行里检测，避免出错
        for cell in row:
            # print(cell.value)
            if ('姓名' in str(cell.value)) or ('名字' in str(cell.value)):
                biaotou_row = cell.row
                name_col = cell.column
            if ('专业班次' in str(cell.value)) or ('班级' in str(cell.value)) or ('专业' in str(cell.value)):
                class_col = cell.column
            if ('推荐支部' in str(cell.value)) or ('支部' in str(cell.value)):
                zhibu_col = cell.column

    for i in range(minrow + 3, maxrow + 1):
        cell1 = worksheet.cell(i, name_col).value
        cell2 = worksheet.cell(i, class_col).value
        cell3 = worksheet.cell(i, zhibu_col).value
        print([cell1, cell2, cell3], end=" ")
        data.append([cell1, cell2, cell3])
    # 使用filter()函数，删除列表中的None值
    data = list(filter(None, data))
    print(data)

    # data = pd.read_excel("互联网/参考 学员花名册7.xlsx",
    #                      sheet_name='Sheet1', usecols=[1, 8, 9])
    # # # 首先将pandas读取的数据转化为array
    # # data_array = np.array(data)
    # # # 然后转化为list形式
    # # data_list =data_array.tolist()
    # # print(list(data_list))
    # # data = list(data_list)[2:]
    # # # print(data[2:])
    # data = data[2:]
    # data = data.values.tolist()
    result = []
    need_change = []

    all_dic = dict()
    all_dic_up = dict()
    for i in range(len(data)):
        if '研' not in data[i][2]:
            if data[i][2] in all_dic.keys():
                all_dic.get(data[i][2]).append([data[i][0], data[i][1], data[i][2]])
            else:
                all_dic[data[i][2]] = [[data[i][0], data[i][1], data[i][2]]]
        else:
            if data[i][2] in all_dic_up.keys():
                all_dic_up.get(data[i][2]).append([data[i][0], data[i][1], data[i][2]])
            else:
                all_dic_up[data[i][2]] = [[data[i][0], data[i][1], data[i][2]]]
    # print(all_dic)  # {'工信支部': [['韩冰凌', '工管181班'], ['邱爽', '工管181班'],……、
    # print(all_dic_up) # {'研二支部': [['罗琼', '会计'], ['陈悦', '会计'], ['朱明慧', '会计'], ['蒋露曦', '会计'], ['刘起林', '会计'], ['张莎', '法律（非法学）']]}
    # print(len(all_dic))
    for i in all_dic:
        if type(all_dic[i]) is not list:
            continue
        if len(all_dic[i]) > 25:
            need_change.append(i)
        if 20 <= len(all_dic[i]) <= 25:
            result.append(all_dic[i])
            all_dic[i] = 'over'
        # print(len(all_dic[i]))
        for j in all_dic:
            if i == j or type(all_dic[j]) is not list or type(all_dic[i]) is not list:
                continue
            if i != j:
                if len(all_dic[i]) + len(all_dic[j]) <= 25:
                    all_dic[i].extend(all_dic[j])  # 注意该函数没有返回值
                    result.append(all_dic[i])
                    all_dic[i] = 'over'
                    all_dic[j] = 'over'

    for i in need_change:
        all_dic[i + '_one'] = 'over'
        result.append(all_dic[i][:20])
        all_dic[i + '_two'] = all_dic[i][20:]
        all_dic.pop(i)
    need_change = []
    for i in all_dic:
        if type(all_dic[i]) is not list:
            continue
        if len(all_dic[i]) > 25:
            need_change.append(i)
    if len(need_change) != 0:
        for i in need_change:
            all_dic[i + '_one'] = 'over'
            result.append(all_dic[i][:20])
            all_dic[i + '_two'] = all_dic[i][20:]
            all_dic.pop(i)

    for i in all_dic:
        if type(all_dic[i]) is not list:
            continue
        if len(all_dic[i]) > 25:
            need_change.append(i)
        if 20 <= len(all_dic[i]) <= 25:
            result.append(all_dic[i])
            all_dic[i] = 'over'
        # print(len(all_dic[i]))
        for j in all_dic:
            if i == j or type(all_dic[j]) is not list or type(all_dic[i]) is not list:
                continue
            if i != j:
                if 20 <= len(all_dic[i]) + len(all_dic[j]) <= 25:
                    all_dic[i].extend(all_dic[j])  # 注意该函数没有返回值
                    result.append(all_dic[i])
                    all_dic[i] = 'over'
                    all_dic[j] = 'over'
    length = 100
    min_length = ''
    for i in all_dic:
        if type(all_dic[i]) is list and len(all_dic[i]) < length:
            min_length = i
            length = len(all_dic[i])
    for i in all_dic:
        if type(all_dic[i]) is list and i != min_length and len(all_dic[min_length]) + len(all_dic[i]) <= 25:
            all_dic[i].extend(all_dic[min_length])
            result.append(all_dic[i])
            break
        if type(all_dic[i]) is list and i != min_length:
            all_dic[i].extend(all_dic[min_length][0:23 - len(all_dic[i])])
            result.append(all_dic[i])
            all_dic[min_length] = all_dic[min_length][23 - len(all_dic[i]):]

    ################################################################################
    need_change_up = []
    all = 0
    for i in all_dic_up:
        all += len(all_dic_up[i])
    tem = []
    if all <= 25:
        for i in all_dic_up:
            tem.extend(all_dic_up[i])
        result.append(tem)
    for i in all_dic_up:
        if type(all_dic_up[i]) is not list:
            continue
        if len(all_dic_up[i]) > 25:
            need_change_up.append(i)
        if 20 <= len(all_dic_up[i]) <= 25:
            result.append(all_dic_up[i])
            all_dic_up[i] = 'over'
        # print(len(all_dic_up[i]))
        for j in all_dic_up:
            if i == j or type(all_dic_up[j]) is not list or type(all_dic_up[i]) is not list:
                continue
            if i != j:
                if len(all_dic_up[i]) + len(all_dic_up[j]) <= 25:
                    all_dic_up[i].extend(all_dic_up[j])  # 注意该函数没有返回值
                    result.append(all_dic_up[i])
                    all_dic_up[i] = 'over'
                    all_dic_up[j] = 'over'

    for i in need_change_up:
        all_dic_up[i + '_one'] = 'over'
        result.append(all_dic_up[i][:20])
        all_dic_up[i + '_two'] = all_dic_up[i][20:]
        all_dic_up.pop(i)

    for i in all_dic_up:
        if type(all_dic_up[i]) is not list:
            continue
        if len(all_dic_up[i]) > 25:
            need_change_up.append(i)
        if 20 <= len(all_dic_up[i]) <= 25:
            result.append(all_dic_up[i])
            all_dic_up[i] = 'over'
        # print(len(all_dic_up[i]))
        for j in all_dic_up:
            if i == j or type(all_dic_up[j]) is not list or type(all_dic_up[i]) is not list:
                continue
            if i != j:
                if 20 <= len(all_dic_up[i]) + len(all_dic_up[j]) <= 25:
                    all_dic_up[i].extend(all_dic_up[j])  # 注意该函数没有返回值
                    result.append(all_dic_up[i])
                    all_dic_up[i] = 'over'
                    all_dic_up[j] = 'over'
    length = 100
    min_length = ''
    for i in all_dic_up:
        if type(all_dic_up[i]) is list and len(all_dic_up[i]) < length:
            min_length = i
            length = len(all_dic_up[i])
    for i in all_dic_up:
        if type(all_dic_up[i]) is list and i != min_length and len(all_dic_up[min_length]) + len(all_dic_up[i]) <= 25:
            all_dic_up[i].extend(all_dic_up[min_length])
            result.append(all_dic_up[i])
            break
        if type(all_dic_up[i]) is list and i != min_length:
            all_dic_up[i].extend(all_dic_up[min_length][0:23 - len(all_dic_up[i])])
            result.append(all_dic_up[i])
            all_dic_up[min_length] = all_dic_up[min_length][23 - len(all_dic_up[i]):]
    ##################################################################################
    # result_df = pd.DataFrame(result)
    print(result)  # [[['卢慧珍', '电商191班'], ['覃小梅', '电商191班'], ……
    # 改为：[[['卢慧珍', '电商191班','电物支部'], ['覃小梅', '电商191班','电物支部'], …… #####################################

    workbook = openpyxl.load_workbook("模板1 分组名单表.xlsx")

    # # 字体对象
    font = Font(name=u'宋体', bold=False, italic=False, size=12)  # bold是否加粗, italic是否斜体
    align = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)
    # # 边框
    # thin = Side(border_style="thin", color="000000")  # 边框样式，颜色
    # border = Border(left=thin, right=thin, top=thin, bottom=thin)  # 边框的位置
    # worksheet_check.cell(len(list_name) + 2 + 1, 1).border = border

    for i in range(len(result)):
        worksheet = workbook.worksheets[i]
        count = 1
        for j in result[i]:
            # print(j[1],j[0]) # 电商191班 卢慧珍
            worksheet.cell(7 + count, 2).value = str(j[2])  ################################################# 写入支部名字
            worksheet.cell(7 + count, 3).value = str(j[0])
            worksheet.cell(7 + count, 4).value = str(j[1])
            worksheet.row_dimensions[7 + count].height = 20  # 调整行高

            count += 1

        # 获取四个区域
        max_row = worksheet.max_row  # 获得最大行数
        max_column = worksheet.max_column  # 获得最大列数
        min_row = worksheet.min_row
        min_column = worksheet.min_column

        # 给区域设置设置框线
        for row in tuple(worksheet[min_row + 8:max_row]):
            for cell in row:
                cell.border = my_border('thin', 'thin', 'thin', 'thin')
                # 设置单元格对齐方式 Alignment(horizontal=水平对齐模式,vertical=垂直对齐模式,text_rotation=旋转角度,wrap_text=是否自动换行)
                cell.alignment = align
                cell.font = font

    workbook.save(save)

for i in range(6):
    fenzu_smart('互联网/参考 学员花名册{}.xlsx'.format(i+5), '互联网/ 分组名单表{}.xlsx'.format(i+5))