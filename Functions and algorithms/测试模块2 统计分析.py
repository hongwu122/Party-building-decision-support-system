import tkintertable
import tkinter
from tkinter import messagebox
from tkinter.filedialog import *
from tkinter import ttk
from tkinter.ttk import *
from tkinter import scrolledtext
# from ttkbootstrap import Style as ttkStyle
from tkintertable import TableCanvas, TableModel
import openpyxl

class VerticalScrolledFrame(Frame):
    """A pure Tkinter scrollable frame that actually works!
    * Use the 'interior' attribute to place widgets inside the scrollable frame
    * Construct and pack/place/grid normally
    * This frame only allows vertical scrolling

    """
    def __init__(self, parent, *args, **kw):
        Frame.__init__(self, parent, *args, **kw)

        # create a canvas object and a vertical scrollbar for scrolling it
        vscrollbar = Scrollbar(self, orient=VERTICAL)
        vscrollbar.pack(fill=Y, side=RIGHT, expand=FALSE)
        canvas = Canvas(self, bd=0, highlightthickness=0,
                        yscrollcommand=vscrollbar.set)
        canvas.pack(side=LEFT, fill=BOTH, expand=TRUE)
        vscrollbar.config(command=canvas.yview)

        # reset the view
        canvas.xview_moveto(0)
        canvas.yview_moveto(0)

        # create a frame inside the canvas which will be scrolled with it
        self.interior = interior = Frame(canvas)
        interior_id = canvas.create_window(0, 0, window=interior,
                                           anchor=NW)

        # track changes to the canvas and frame width and sync them,
        # also updating the scrollbar
        def _configure_interior(event):
            # update the scrollbars to match the size of the inner frame
            size = (interior.winfo_reqwidth(), interior.winfo_reqheight())
            # print("size",size)
            canvas.config(scrollregion="0 0 %s %s" % size)
            if interior.winfo_reqwidth() != canvas.winfo_width():
                # update the canvas's width to fit the inner frame
                canvas.config(width=interior.winfo_reqwidth())
        interior.bind('<Configure>', _configure_interior)

        def _configure_canvas(event):
            if interior.winfo_reqwidth() != canvas.winfo_width():
                # update the inner frame's width to fill the canvas
                canvas.itemconfigure(interior_id, width=canvas.winfo_width())
        canvas.bind('<Configure>', _configure_canvas)


window = Tk()
# 窗口的大小,前两、个参数是：宽、高，后面的参数是坐标
# # 设置窗口居中
width =  650  # 1300
height =  690  # 830
screenwidth = window.winfo_screenwidth()
screenheight = window.winfo_screenheight()
alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth-width)/2, (screenheight-height-100)/2)
window.geometry(alignstr)
# 禁止窗口的拉伸
# window.resizable(0, 0)  # == (Flase,Flase)
# 窗口置顶
# window.attributes("-topmost", 1)  # 1==True 处于顶层
# 标题
window.title('党建决策支持系统')

try:
    window.iconbitmap('mould\ico.ico')
except:pass
# 设置背景颜色
window['background']='gray97'
# messagebox.showinfo('TO YOU', message='毛主席万岁')

# 分区
tabControl = ttk.Notebook(window)       # Create Tab Control 总菜单

# tab6_8 =  ttk.Frame(tabControl)
# tabControl.add(tab6_8, text=' 请示批复备案 ')    # Add the tab 二级菜单

tab18 = ttk.Frame(tabControl)
tabControl.add(tab18, text=' 统计分析 ')

tabControl.pack(expand=1, fill="both")  # Pack to make visible

####################################################################################################################
import numpy as np
import matplotlib.pyplot as plt
# 支持中文
plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure

def statistics_Toplevel():
    statistics_Toplevel = Toplevel()
    screenwidth = statistics_Toplevel.winfo_screenwidth()
    screenheight = statistics_Toplevel.winfo_screenheight()
    width = screenwidth-100  # 1300
    height = screenheight-100  # 830
    statistics_Toplevel.geometry('%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height - 50) / 2))  # 设置主窗口大小

    # 测试 统计分析
    mighty18_1 = ttk.LabelFrame(statistics_Toplevel, text='') # ,style='Frame1.TFrame'
    mighty18_1.place(x=0,y=0,width=width,height=400) # 64
    # table = TableCanvas(mighty18_1)
    # table.show()
    model = TableModel()
    # table = TableCanvas(mighty18_1, model=model)
    table = TableCanvas(mighty18_1, model,
                cellwidth=120, cellbackgr='#e3f698',
                thefont=('Arial',12),rowheight=24, rowheaderwidth=30,
                rowselectedcolor='yellow', editable=True,rows=100, cols=100,)
    try:
        table.importCSV('mould/statistics.csv')
    except:pass
    table.show()
    button = Button(master=statistics_Toplevel, text="保存", command= lambda :save_table(table,"test.txt") )
    # 按钮放在下边
    button.place(x=170, y=415)
    statistics_Toplevel.mainloop()

def save_table(table,filename):
    table.save(filename)
    messagebox.showinfo("提示","文件：{}，保存成功！".format(filename))

def keshihua():
    win = tkinter.Toplevel()
    win.title("党员发展情况可视化")
    # f = Figure(figsize=(5, 4), dpi=100)
    # a = f.add_subplot(111)  # 添加子图:1行1列第1个
    f = plt.figure()

    year = ['2019年', '2020年', '2021年', '2022年']
    # jj = np.array([672, 150, 441, 364])
    # fz = np.array([195, 178, 428, 159])
    # yb = np.array([138, 167, 406, 153])
    # dy = np.array([147, 138, 134, 360])
    jj = np.array([672, 400, 441, 364])
    fz = np.array([195, 178, 428, 159])
    yb = np.array([138, 167, 406, 153])
    dy = np.array([147, 138, 134, 360])
    # 此处的 _ 下划线表示将循环取到的值放弃，只得到[0,1,2,3,4]
    # ind = [x for x, _ in enumerate(year)]
    #绘制堆叠图
    plt.bar(year, jj, width=0.5, label='积极分子', color='gold') # , bottom=
    plt.bar(year, fz, width=0.5, label='发展对象', color='orange')
    plt.bar(year, yb, width=0.5, label='预备党员', color='#CD853F')
    plt.bar(year, dy, width=0.5, label='转正党员', color='red')

    #设置坐标轴
    # plt.xticks(ind, year)
    plt.ylabel("人数")
    plt.xlabel("年份")
    plt.legend(loc="upper right")
    plt.title("党员发展堆叠图")
    # plt.show()


    # 将绘制的图形显示到tkinter:创建属于root的canvas画布,并将图f置于画布上
    canvas = FigureCanvasTkAgg(f, master=win)
    canvas.draw()  # 注意show方法已经过时了,这里改用draw
    plt.draw()

    canvas.get_tk_widget().pack(side=tkinter.TOP,  # 上对齐
                                fill=tkinter.BOTH,  # 填充方式
                                expand=tkinter.YES)  # 随窗口大小调整而调整

    # matplotlib的导航工具栏显示上来(默认是不会显示它的)
    toolbar = NavigationToolbar2Tk(canvas, win)
    toolbar.update()
    canvas._tkcanvas.pack(side=tkinter.TOP,  # get_tk_widget()得到的就是_tkcanvas
                          fill=tkinter.BOTH,
                          expand=tkinter.YES)

def keshihua2():
    win = tkinter.Toplevel()
    win.title("数据统计可视化")
    # f = Figure(figsize=(5, 4), dpi=100)
    # a = f.add_subplot(111)  # 添加子图:1行1列第1个
    f = plt.figure()

    labels = ['积极分子', '发展对象','预备党员', '转正党员']
    a = [672, 195, 138, 147]
    b = [400, 178, 167, 138]
    c = [441, 428, 406, 134]
    d = [364, 159, 153, 360]
    year = ['2019年', '2020年', '2021年', '2022年']

    x = np.arange(len(labels))  # 标签位置
    width = 0.1  # 柱状图的宽度，可以根据自己的需求和审美来改

    # fig, ax = plt.subplots()
    rects1 = plt.bar(x - width * 2, a, width, label=labels[0])
    rects2 = plt.bar(x - width + 0.01, b, width, label=labels[1])
    rects3 = plt.bar(x + 0.02, c, width, label=labels[2])
    rects4 = plt.bar(x + width + 0.03, d, width, label=labels[3])

    # 为y轴、标题和x轴等添加一些文本。
    plt.ylabel('人数', fontsize=16)
    plt.xlabel('年份', fontsize=16)
    plt.title('近四年党员发展人数柱状图')
    plt.xticks(x,year)
    # plt.xticklabels(labels)
    plt.legend()

    def autolabel(rects):
        """在*rects*中的每个柱状条上方附加一个文本标签，显示其高度"""
        for rect in rects:
            height = rect.get_height()
            plt.annotate('{}'.format(height),
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(0, 3),  # 3点垂直偏移
                        textcoords="offset points",
                        ha='center', va='bottom')

    autolabel(rects1)
    autolabel(rects2)
    autolabel(rects3)
    autolabel(rects4)

    # fig.tight_layout()


    # 将绘制的图形显示到tkinter:创建属于root的canvas画布,并将图f置于画布上
    canvas = FigureCanvasTkAgg(f, master=win)
    canvas.draw()  # 注意show方法已经过时了,这里改用draw
    plt.draw()

    canvas.get_tk_widget().pack(side=tkinter.TOP,  # 上对齐
                                fill=tkinter.BOTH,  # 填充方式
                                expand=tkinter.YES)  # 随窗口大小调整而调整

    # matplotlib的导航工具栏显示上来(默认是不会显示它的)
    toolbar = NavigationToolbar2Tk(canvas, win)
    toolbar.update()
    canvas._tkcanvas.pack(side=tkinter.TOP,  # get_tk_widget()得到的就是_tkcanvas
                          fill=tkinter.BOTH,
                          expand=tkinter.YES)


def statistics_vision(filename_i):
    filename = 'mould//csv//' + filename_i + ".csv"
    # try:
    table.importCSV(filename) # 只支持gbk的编码文件导入
    table.show()
    print(filename)
    # except:
    #     messagebox.showinfo("提示","文件：{}，导入错误！".format(filename))

import pandas as pd
import os
def excel2csv(excel_file):
    # 打开excel文件
    workbook = openpyxl.load_workbook(excel_file)
    # 获取所有sheet名字
    sheet_names = workbook.sheetnames
    for worksheet_name in sheet_names:
        # 遍历每个sheet并用Pandas读取
        data_xls = pd.read_excel(excel_file, worksheet_name, index_col=None)
        # 获取excel当前目录
        dir_path = os.path.abspath(os.path.dirname(excel_file))
        # 转换成csv并保存到excel所在目录下的csv文件夹中
        csv_path = 'mould\\csv\\'  # dir_path +
        if not os.path.exists(csv_path):
            os.mkdir(csv_path)
        data_xls.to_csv(csv_path + worksheet_name + '.csv', index=None, encoding='gbk') # utf-8-sig
    print("生成csv成功！")

def zairu():
    filename = "mould//模板7 请示批复一览表.xlsx"
    try:
        excel2csv(filename)
    except:
        messagebox.showinfo("提示", "文件：{}，载入失败，请检查文件位置是否存在！".format(filename))


# 创建一个按钮,并把上面那个函数绑定过来
button = Button(master=tab18, text="载入表格", command=zairu)
button.place(x = 140, y = 415) # 按钮放在下边

button = Button(master=tab18, text="可视化", command=keshihua)
button.place(x = 240, y = 415)

button = Button(master=tab18, text="数据分析", command=keshihua2)
button.place(x = 340, y = 415)

button = Button(master=tab18, text="全屏显示", command=statistics_Toplevel)
button.place(x = 440, y = 415)

button = Button(master=tab18, text="保存", command=lambda: save_table(table, "test.txt"))
button.place(x = 540, y = 415)


# 选择学期滚动条
mighty18_2 = VerticalScrolledFrame(tab18)
mighty18_2.place(x = 10, y = 40, height=400)
label = Label(master=tab18, text="请选择学期：")
label.place(x = 10, y = 10)
buttons = []


try:
    workbook = openpyxl.load_workbook("mould//模板7 请示批复一览表.xlsx")
    statistics_time = workbook.sheetnames
    print(workbook.sheetnames)
except:
    statistics_time = ['2018下', '2019上', '2019下', '2020上', '2020下', '2021上', '2021中', '2021下', '2022上', '2022下', '2023上', '2023下', '2024上', '2024下']

for i in statistics_time:
    buttons.append(Button(master=mighty18_2.interior, text = i , command = lambda arg=i:statistics_vision(arg))) # command = lambda : statistics_vision(i)
    # 上述代码实现，导入多个按钮的动态传参数效果
    buttons[-1].pack()
# print(buttons)



# 表格展示区
mighty18_2 = ttk.LabelFrame(tab18, text='') # ,style='Frame1.TFrame'
mighty18_2.place(x=120,y=10,width=650-120-10,height=400) # 64
# table = TableCanvas(mighty18_1)
# table.show()
model = TableModel()
# table = TableCanvas(mighty18_1, model=model)
table = TableCanvas(mighty18_2, model,
            cellwidth=120, cellbackgr='#e3f698',
            thefont=('Arial',12),rowheight=24, rowheaderwidth=30,
            rowselectedcolor='yellow', editable=True,rows=100, cols=100,)























window.mainloop()










# # 测试 统计分析
# mighty18_1 = ttk.LabelFrame(tab18, text='') # ,style='Frame1.TFrame'
# mighty18_1.place(x=0,y=0,width=1800,height=400) # 64
# # table = TableCanvas(mighty18_1)
# # table.show()
# model = TableModel()
# # table = TableCanvas(mighty18_1, model=model)
# table = TableCanvas(mighty18_1, model,
#             cellwidth=120, cellbackgr='#e3f698',
#             thefont=('Arial',12),rowheight=24, rowheaderwidth=30,
#             rowselectedcolor='yellow', editable=True,rows=100, cols=100,)
# try:
#     table.importCSV('mould/statistics.csv')
# except:pass
# table.show()
#
#
# import numpy as np
# import matplotlib.pyplot as plt
# # 支持中文
# plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
# plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号
# import tkinter
# from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
# from matplotlib.backend_bases import key_press_handler
# from matplotlib.figure import Figure
#
# def keshihua():
#     win = tkinter.Toplevel()
#     win.title("党员发展情况可视化")
#     # f = Figure(figsize=(5, 4), dpi=100)
#     # a = f.add_subplot(111)  # 添加子图:1行1列第1个
#     f = plt.figure()
#
#     year = ['2019年', '2020年', '2021年', '2022年']
#     # jj = np.array([672, 150, 441, 364])
#     # fz = np.array([195, 178, 428, 159])
#     # yb = np.array([138, 167, 406, 153])
#     # dy = np.array([147, 138, 134, 360])
#     jj = np.array([672, 400, 441, 364])
#     fz = np.array([195, 178, 428, 159])
#     yb = np.array([138, 167, 406, 153])
#     dy = np.array([147, 138, 134, 360])
#     # 此处的 _ 下划线表示将循环取到的值放弃，只得到[0,1,2,3,4]
#     # ind = [x for x, _ in enumerate(year)]
#     #绘制堆叠图
#     plt.bar(year, jj, width=0.5, label='积极分子', color='gold') # , bottom=
#     plt.bar(year, fz, width=0.5, label='发展对象', color='orange')
#     plt.bar(year, yb, width=0.5, label='预备党员', color='#CD853F')
#     plt.bar(year, dy, width=0.5, label='转正党员', color='red')
#
#     #设置坐标轴
#     # plt.xticks(ind, year)
#     plt.ylabel("人数")
#     plt.xlabel("年份")
#     plt.legend(loc="upper right")
#     plt.title("党员发展堆叠图")
#     # plt.show()
#
#
#     # 将绘制的图形显示到tkinter:创建属于root的canvas画布,并将图f置于画布上
#     canvas = FigureCanvasTkAgg(f, master=win)
#     canvas.draw()  # 注意show方法已经过时了,这里改用draw
#     plt.draw()
#
#     canvas.get_tk_widget().pack(side=tkinter.TOP,  # 上对齐
#                                 fill=tkinter.BOTH,  # 填充方式
#                                 expand=tkinter.YES)  # 随窗口大小调整而调整
#
#     # matplotlib的导航工具栏显示上来(默认是不会显示它的)
#     toolbar = NavigationToolbar2Tk(canvas, win)
#     toolbar.update()
#     canvas._tkcanvas.pack(side=tkinter.TOP,  # get_tk_widget()得到的就是_tkcanvas
#                           fill=tkinter.BOTH,
#                           expand=tkinter.YES)
#
# def keshihua2():
#     win = tkinter.Toplevel()
#     win.title("数据统计可视化")
#     # f = Figure(figsize=(5, 4), dpi=100)
#     # a = f.add_subplot(111)  # 添加子图:1行1列第1个
#     f = plt.figure()
#
#     labels = ['积极分子', '发展对象','预备党员', '转正党员']
#     a = [672, 195, 138, 147]
#     b = [400, 178, 167, 138]
#     c = [441, 428, 406, 134]
#     d = [364, 159, 153, 360]
#     year = ['2019年', '2020年', '2021年', '2022年']
#
#     x = np.arange(len(labels))  # 标签位置
#     width = 0.1  # 柱状图的宽度，可以根据自己的需求和审美来改
#
#     # fig, ax = plt.subplots()
#     rects1 = plt.bar(x - width * 2, a, width, label=labels[0])
#     rects2 = plt.bar(x - width + 0.01, b, width, label=labels[1])
#     rects3 = plt.bar(x + 0.02, c, width, label=labels[2])
#     rects4 = plt.bar(x + width + 0.03, d, width, label=labels[3])
#
#     # 为y轴、标题和x轴等添加一些文本。
#     plt.ylabel('人数', fontsize=16)
#     plt.xlabel('年份', fontsize=16)
#     plt.title('近四年党员发展人数柱状图')
#     plt.xticks(x,year)
#     # plt.xticklabels(labels)
#     plt.legend()
#
#     def autolabel(rects):
#         """在*rects*中的每个柱状条上方附加一个文本标签，显示其高度"""
#         for rect in rects:
#             height = rect.get_height()
#             plt.annotate('{}'.format(height),
#                         xy=(rect.get_x() + rect.get_width() / 2, height),
#                         xytext=(0, 3),  # 3点垂直偏移
#                         textcoords="offset points",
#                         ha='center', va='bottom')
#
#     autolabel(rects1)
#     autolabel(rects2)
#     autolabel(rects3)
#     autolabel(rects4)
#
#     # fig.tight_layout()
#
#
#     # 将绘制的图形显示到tkinter:创建属于root的canvas画布,并将图f置于画布上
#     canvas = FigureCanvasTkAgg(f, master=win)
#     canvas.draw()  # 注意show方法已经过时了,这里改用draw
#     plt.draw()
#
#     canvas.get_tk_widget().pack(side=tkinter.TOP,  # 上对齐
#                                 fill=tkinter.BOTH,  # 填充方式
#                                 expand=tkinter.YES)  # 随窗口大小调整而调整
#
#     # matplotlib的导航工具栏显示上来(默认是不会显示它的)
#     toolbar = NavigationToolbar2Tk(canvas, win)
#     toolbar.update()
#     canvas._tkcanvas.pack(side=tkinter.TOP,  # get_tk_widget()得到的就是_tkcanvas
#                           fill=tkinter.BOTH,
#                           expand=tkinter.YES)
#
# # 创建一个按钮,并把上面那个函数绑定过来
# button = Button(master=tab18, text="可视化", command=keshihua)
# # 按钮放在下边
# button.place(x = 170, y = 415)
# # 创建一个按钮,并把上面那个函数绑定过来
# button = Button(master=tab18, text="数据分析", command=keshihua2)
# # 按钮放在下边
# button.place(x = 400, y = 415)
#
#
#
#
# # data=table.model.data
# # table.model.setValueAt(123,3,4) # 插入
# # table.model.setValueAt('hahaha',2,4)
# # table.model.setValueAt(repr(type('haha')),1,4)
#
# # import numpy as np
# # import matplotlib
# # import matplotlib.pyplot as plt
# # from matplotlib.pylab import mpl
# # from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
# #
# # class Application(Frame):
# #     """一个经典的GUI写法"""
# #
# #     def __init__(self, master=None):
# #         '''初始化方法'''
# #         super().__init__(master)  # 调用父类的初始化方法
# #         self.master = master
# #         self.pack(side=TOP, fill=BOTH, expand=1)  # 此处填充父窗体
# #         self.create_matplotlib_zhuzhuantu()
# #         # self.create_matplotlib_quxiantu()
# #         self.createWidget(self.figure)
# #
# #     def createWidget(self, figure):
# #         """创建组件"""
# #         # self.label = Label(self, text='党员信息统计与管理')
# #         # self.label.pack()
# #         # 创建画布
# #         self.canvas = FigureCanvasTkAgg(figure, self)
# #         self.canvas.draw()
# #         self.canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)
# #         # 把matplotlib绘制图形的导航工具栏显示到tkinter窗口上
# #         # toolbar = NavigationToolbar2Tk(self.canvas, self)
# #         # toolbar.update()
# #         # self.canvas._tkcanvas.pack(side=TOP, fill=BOTH, expand=1)
# #         # self.button = Button(master=self, text="退出", command=quit)
# #         # # 按钮放在下边
# #         # self.button.pack(side=BOTTOM)
# #
# #     def create_matplotlib_zhuzhuantu(self):
# #         """创建绘图对象"""
# #         # 设置中文显示字体
# #         mpl.rcParams['font.sans-serif'] = ['SimHei']  # 中文显示
# #         mpl.rcParams['axes.unicode_minus'] = False  # 负号显示
# #         # 创建绘图对象f figsize的单位是英寸 像素 = 英寸*分辨率
# #         self.figure = plt.figure(num=2, figsize=(7, 4), dpi=80,  frameon=True) # ,edgecolor='red', facecolor="lightskyblue"
# #         self.figure.text(4, 9, '2020支部新发展党员数量图')  # 设置显示的文本
# #         x = np.arange(12)
# #         y = np.random.uniform(0, 14, 12) # * (1 - x / float(12))
# #         loc = zip(x, y)  # 将x, y 两两配对
# #         plt.ylim(0, 20)  # 设置y轴的范围
# #         plt.bar(x, y, facecolor='red', edgecolor='black')  # 绘制柱状图(填充颜色绿色，边框黑色)
# #         for x, y in loc:
# #             plt.text(x, y, '{}'.format(round(y,0)), ha='center', va='bottom')  # 保留小数点2位
# #     # def create_matplotlib_quxiantu(self):
# #     #     """创建绘图对象"""
# #     #     # 设置中文显示字体
# #     #     mpl.rcParams['font.sans-serif'] = ['SimHei']  # 中文显示
# #     #     mpl.rcParams['axes.unicode_minus'] = False  # 负号显示
# #     #     # 创建绘图对象f figsize的单位是英寸 像素 = 英寸*分辨率
# #     #     self.figure = plt.figure(num=2, figsize=(7, 4), dpi=80, facecolor="gold", edgecolor='green', frameon=True)
# #     #     # 创建一副子图
# #     #     fig1 = plt.subplot(1, 1, 1)  # 三个参数，依次是：行，列，当前索引
# #     #     # 创建数据源：x轴是等间距的一组数
# #     #     x = np.arange(-2 * np.pi, 2 * np.pi, 0.1)
# #     #     y1 = np.sin(x)
# #     #     y2 = np.cos(x)
# #     #
# #     #     line1 = fig1.plot(x, y1, color='red', linewidth=2, label='y=sin(x)', linestyle='--')  # 画第一条线
# #     #     line2 = fig1.plot(x, y2, color='green', label='y=cos(x)')
# #     #     plt.setp(line2, linewidth=1, linestyle='-', alpha=0.7)  # 华第二条线 color='',
# #     #
# #     #     fig1.set_title("数学曲线图", loc='center', pad=20, fontsize='xx-large', color='red')  # 设置标题
# #     #     # line1.set_label("正弦曲线")  # 确定图例
# #     #     # 定义legend 重新定义了一次label
# #     #     fig1.legend(['正弦', '余弦'], loc='lower right', facecolor='orange', frameon=True, shadow=True, framealpha=0.7)
# #     #     # ,fontsize='xx-large'
# #     #     fig1.set_xlabel('(x)横坐标')  # 确定坐标轴标题
# #     #     fig1.set_ylabel("(y)纵坐标")
# #     #     fig1.set_yticks([-1, -1 / 2, 0, 1 / 2, 1])  # 设置坐标轴刻度
# #     #     fig1.grid(which='major', axis='x', color='gray', linestyle='-', linewidth=0.5, alpha=0.2)  # 设置网格
# #
# #     def destroy(self):
# #         """重写destroy方法"""
# #         super().destroy()
# #         quit()
# # app = Application(master=tab17)

from docx.table import _Cell
print(_Cell)