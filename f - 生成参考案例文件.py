
from openpyxl.styles import Color, Font, Alignment, Border, Side, PatternFill, colors
from openpyxl.utils import get_column_letter
import openpyxl
from faker import Faker
import random
import os

# 定义边框样式
def my_border(t_border, b_border, l_border, r_border):
    border = Border(top=Side(border_style=t_border, color=colors.BLACK),
                    bottom=Side(border_style=b_border, color=colors.BLACK),
                    left=Side(border_style=l_border, color=colors.BLACK),
                    right=Side(border_style=r_border, color=colors.BLACK))
    return border

def getBirthday(year_min,year_max):
        # 随机生成年月日
        year = random.randint(year_min, year_max)
        month = random.randint(1, 12)
        # 判断每个月有多少天随机生成日
        if year % 4 == 0:
                if month in (1, 3, 5, 7, 8, 10, 12):
                        day = random.randint(1, 31)
                elif month in (4, 6, 9, 11):
                        day = random.randint(1, 30)
                else:
                        day = random.randint(1, 29)
        else:
                if month in (1, 3, 5, 7, 8, 10, 12):
                        day = random.randint(1, 31)
                elif month in (4, 6, 9, 11):
                        day = random.randint(1, 30)
                else:
                        day = random.randint(1, 28)
        # 小于10的月份前面加0
        if month < 10:
                month = '0' + str(month)
        if day < 10:
                day = '0' + str(day)
        birthday = str(year) + str(month) + str(day)
        return birthday

def test_file_book():
        fake = Faker(locale='zh_CN')

        zhibu_list = ['电物支部','工信支部','会计一支部','会计二支部','国贸支部','经济支部','研一支部','研二支部','法学支部','人营支部']
        zhuanye_list = [["电商","物流"], ["工管","信管","企管"], ["会计","ACCA"], ["会计","ACCA"], ["国贸"], ["经济"],
                        ["工商管理","公共管理","应用经济学"], ["会计","MBA","法律（非法学）"], ["法学"], ["人管","营销"]]
        # 遍历写入各个支部的学员册
        for zb in range(len(zhibu_list)):
                # 新建表
                workbook = openpyxl.load_workbook('mould\模板0 学员花名册.xlsx')  # 打开模板表
                worksheet = workbook.worksheets[0]

                # 删除红行
                worksheet.delete_rows(4)

                # 该支部学员册文件的人数上线，随机
                count = random.randint(10, 30)
                # print(count)
                # 写入学员册里面的人
                for row in range(4,count):
                        # 学号  姓名	性别	出生年月	民族	籍贯	所属院系	年级	专业班次	现任学生干部职务 首次递交入党申请书时间  确认为入党积极分子时间  是否为团员  备注
                        # 随机学号
                        worksheet.cell(row, 1, fake.random_int(min=20150000000,max=20210000000))

                        # 随机名字和性别
                        if random.randint(0,1)==1:
                                worksheet.cell(row, 2, fake.name_female())
                                worksheet.cell(row, 3, '男')
                        else:
                                worksheet.cell(row, 2, fake.name_male())
                                worksheet.cell(row, 3, '女')

                        # 随机出生日期
                        worksheet.cell(row, 4, getBirthday(year_min=1996,year_max=2003) )

                        # 随机民族
                        mz = random.randint(1,10)
                        if mz < 10:
                                worksheet.cell(row, 5, '汉族')
                        if mz == 10:
                                mz_list = ["汉族","蒙古族","回族","藏族","维吾尔族","苗族","彝族","壮族","布依族","朝鲜族","满族","侗族","瑶族","白族","土家族","哈尼族","哈萨克族","傣族","黎族","僳僳族",
                                "佤族","畲族","高山族","拉祜族","水族","东乡族","纳西族","景颇族","柯尔克孜族","土族","达斡尔族","仫佬族","羌族","布朗族","撒拉族","毛南族","仡佬族","锡伯族",
                                 "阿昌族","普米族","塔吉克族","怒族","乌孜别克族","俄罗斯族","鄂温克族","德昂族","保安族","裕固族","京族","塔塔尔族","独龙族","鄂伦春族","赫哲族","门巴族","珞巴族","基诺族"]
                                worksheet.cell(row, 5, random.choice(mz_list) )

                        # 随机籍贯
                        # 这里懒得核对省份城市了，后人可以完善一下
                        provinces = ["北京", "上海", "天津", "重庆","内蒙古", "山西", "河北", "吉林", "江苏", "辽宁", "黑龙江","安徽", "山东", "浙江", "江西", "福建", "湖南", "湖北",
                                "河南", "广东", "广西", "贵州", "海南", "四川", "云南","陕西", "甘肃省", "宁夏", "青海", "新疆", "西藏","台湾", "香港", "澳门"]
                        cities = ["哈尔滨", "长春", "沈阳", "呼和浩特", "石家庄", "乌鲁木齐", "兰州", "西宁", "西安", "银川", "郑州", "济南", "太原","合肥", "武汉", "长沙", "南京", "成都", "贵阳", "昆明", "南宁", "拉萨",
                                "杭州", "南昌", "广州", "福州", "台北", "海口", "郴州", "宁乡", "怀化", "太原", "辛集", "邯郸", "沈阳", "娄底", "兴城", "北镇", "阜新","哈尔滨", "衡阳", "湘西", "张家界", "常德",
                                "六安", "巢湖", "马鞍山", "永安", "宁德", "嘉禾", "荆门", "潜江", "大冶", "宜都", "佛山", "深圳","潮州", "惠州", "汕尾", "东莞", "梧州", "湘潭", "长沙", "株洲", "益阳"]
                        worksheet.cell(row, 6, random.choice(provinces) + random.choice(cities))

                        # yuanxi_list = ["机械工程学院","电气工程学院","核科学技术学院","资源环境与安全工程学院","计算机学院","土木工程学院","建筑学院","松霖设计艺术学院",
                        #                "化学化工学院","数理学院","衡阳医学院","药学院","公共卫生学院","护理学院","经济管理与法学学院","语言文学学院","马克思主义学院",
                        #                "体育学院","国际学院","船山学院","创新创业学院","继续教育学院"]
                        '''# 这里提供本校的一些院系专业供后人开发'''
                        # zhuanye_dict = {"机械工程学院": ("机械设计制造及其自动化", "材料成型及控制工程", "过程装备与控制工程", "测控技术与仪器", "能源与动力工程", "车辆工程"),
                        #                 "电气工程学院": ("电子信息工程", "电子信息科学与技术", "电气工程及其自动化", "自动化", "通信工程", "生物医学工程"),
                        #                 "计算机学院": ("软件工程", "物联网工程", "网络工程", "数字媒体技术", "医学信息工程"),
                        #                 "土木工程学院": ("土木工程", "建筑环境与能源应用工程", "给排水科学与工程", "建筑电气与智能化", "道路桥梁与渡河工程"),
                        #                 "化学化工学院": ("化学工程与工艺", "制药工程", "高分子材料与工程", "无机非金属材料工程"),
                        #                 "核科学技术学院": ("核工程与核技术", "辐射防护与核安全", "核化工与核燃料工程", "核物理"),
                        #                 "资源环境与安全工程学院": ("矿物资源工程", "矿物加工工程", "资源勘查工程", "城市地下空间工程", "安全工程", "环境工程", "环保设备工程"),
                        #                 "松霖设计艺术学院": ("工业设计", "建筑学", "风景园林", "城乡规划", "视觉传达设计", "环境设计", "产品设计", "数字媒体艺术"),
                        #                 "数理学院": ("信息与计算科学"),
                        #                 "衡阳医学院": ("临床医学", "医学检验技术", "医学影像学", "口腔医学", "麻醉学", "儿科学", "生物技术"),
                        #                 "药学院": ("药学", "药物制剂", "预防医学", "卫生检验与检疫"),
                        #                 "护理学院": ("护理学"),
                        #                 "经济管理与法学学院": (
                        #                 "工商管理", "人力资源管理", "市场营销", "会计学", "电子商务", "物流工程", "国际经济与贸易", "经济学", "法学", "信息管理与信息系统"),
                        #                 "语言文学学院": ("英语", "翻译", "日语", "汉语言文学")}
                        '''# 随机院系及专业'''
                        # yuanxi = random.choice(list(zhuanye_dict)) # 随机院系
                        yuanxi = '经济管理与法学学院'
                        zhuanye = random.choice(zhuanye_list[zb])

                        worksheet.cell(row, 7, yuanxi)

                        # 随机年级
                        nianji_list = [x for x in range(15,22)]
                        nianji = str(random.choice(nianji_list))
                        worksheet.cell(row, 8,  '20{}级'.format(nianji)    )

                        # 随机专业班次
                        banji = '{}{}{}班'.format(zhuanye, nianji, random.choice([x for x in range(1,16)]))
                        worksheet.cell(row, 9, banji)

                        # 随机现任学生干部职务
                        zhiwu_list = ['宣传委员','生活委员','班长','团支书','副班长','组织委员','文娱委员','科协委员','心理委员']
                        zw = random.randint(0,1)
                        if zw == 0:
                                worksheet.cell(row, 10, '无')
                        if zw == 1:
                                worksheet.cell(row, 10, banji + random.choice(zhiwu_list)  )

                        # 首次递交入党申请书时间
                        scrd = getBirthday(year_min=2019, year_max=2020)
                        worksheet.cell(row, 11, scrd)

                        # 确认为入党积极分子时间
                        jjfz = getBirthday(year_min=int(scrd[0:4])+1, year_max=int(scrd[:4])+1)
                        worksheet.cell(row, 12,  jjfz)

                        # 是否团员
                        worksheet.cell(row, 13, '是')

                        # QQ
                        worksheet.cell(row, 14, str(fake.phone_number())[:9]   )

                        # 推荐支部
                        worksheet.cell(row, 15, str(zhibu_list[zb])   )

                        # 备注
                        bz = random.randint(1,10)
                        beizhu = ['抗疫志愿者','优秀团干部']
                        if bz == 10:
                                worksheet.cell(row, 16, random.choice(beizhu)   )
                        else:
                                pass
                # 获取四个区域
                max_row = worksheet.max_row  # 获得最大行数
                # max_column = worksheet.max_column  # 获得最大列数
                min_row = worksheet.min_row
                # min_column = worksheet.min_column
                # 给区域设置设置框线
                for row in tuple(worksheet[min_row:max_row]):
                        for cell in row:
                                cell.border = my_border('thin', 'thin', 'thin', 'thin')
                                # 设置单元格对齐方式 Alignment(horizontal=水平对齐模式,vertical=垂直对齐模式,text_rotation=旋转角度,wrap_text=是否自动换行)
                                cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)

                if os.path.exists('参考 各支部学员花名册') is False:
                        print("文件夹不存在")
                        os.mkdir("参考 各支部学员花名册")

                workbook.save('参考 各支部学员花名册/{} 入党积极分子名册（2021上）.xlsx'.format(zhibu_list[zb]))

def test_file_count():
        # # 字典不好搞遍历？？
        # zhibu_dict = {'电物支部':("电商","物流"),'工信支部':("工管","信管","企管"),'会计一支部':("会计","ACCA"),'法学支部':("法学"),
        #                 '会计二支部':("会计","ACCA"),'国贸支部':("国贸"),'经济支部':("经济"),'人营支部':("人管","营销"),
        #               '研一支部':("工商管理","公共管理","应用经济学"),'研二支部':("会计","MBA","法律（非法学）")}
        zhibu_list = ['电物支部','工信支部','会计一支部','会计二支部','国贸支部','经济支部','研一支部','研二支部','法学支部','人营支部']
        zhuanye_list = [["电商","物流"], ["工管","信管","企管"], ["会计","ACCA"], ["会计","ACCA"], ["国贸"], ["经济"],
                        ["工商管理","公共管理","应用经济学"], ["会计","MBA","法律（非法学）"], ["法学"], ["人管","营销"]]

        # 遍历写入每个支部文件
        for zb in range(len(zhibu_list)):
                # 新建表
                workbook = openpyxl.Workbook()
                worksheet = workbook.worksheets[0]

                # 写入表头
                worksheet.cell(1, 1, '支部')
                worksheet.cell(1, 2, '班级')
                worksheet.cell(1, 3, '递交入党申请书人数')

                print('正在遍历支部是：{}'.format(zhibu_list[zb]))
                row = 2 # 从第二行开始写
                # 遍历写入专业
                zhuanye = zhuanye_list[zb]
                nianji = ['17','18','19','20']
                for bj in range(len(zhuanye)):
                        print('正在遍历专业是：{}'.format(zhuanye[bj]))
                        # 随机写入班级   zhuanye_list[zb][bj]==每个专业
                        class_names = []
                        for i in range(random.randint(4,12)): # 设置人数概率
                                class_n = zhuanye[bj] + random.choice(nianji) + str(random.randint(1,5)) + '班'
                                if class_n not in class_names:
                                        worksheet.cell(row, 2, class_n)  # 写入班级全名
                                        worksheet.cell(row, 3, random.randint(0, 12))  # 写入随机的递交入党申请书人数
                                        row = row + 1
                                class_names.append(class_n)


                # print('本次写到了行数：{}'.format(row))

                worksheet.cell(2, 1, zhibu_list[zb])  # 写入支部名字
                worksheet.merge_cells(start_row=2, start_column=1, end_row=row-1, end_column=1) # 合并单元格

                # 获取四个区域
                max_row = worksheet.max_row  # 获得最大行数
                max_column = worksheet.max_column  # 获得最大列数
                min_row = worksheet.min_row
                min_column = worksheet.min_column

                # 给区域设置设置框线
                for r in tuple(worksheet[min_row:max_row]):
                        for cell in r:
                                cell.border = my_border('thin', 'thin', 'thin', 'thin')
                                # 设置单元格对齐方式 Alignment(horizontal=水平对齐模式,vertical=垂直对齐模式,text_rotation=旋转角度,wrap_text=是否自动换行)
                                cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)

                # 区域自动调整列宽
                column_widths = []  # 定义用来获取当前列最大宽度的空列表
                for i, col in enumerate(worksheet.iter_cols(min_col=min_column, max_col=max_column, min_row=min_row, max_row=max_row)):
                        for cell in col:
                                value = cell.value
                                if value is not None:
                                        if isinstance(value, str) is False:
                                                value = str(value)
                                        try:
                                                column_widths[i] = max(column_widths[i], len(value))
                                        except IndexError:
                                                column_widths.append(len(value))
                # print('column_widths', column_widths)  # 得到该列最大的一个单元格的宽度（字符串数量）
                for i, width in enumerate(column_widths):
                        col_name = get_column_letter(min_column + i)  # 获取行字母表头
                        value = column_widths[i] * 2  # 设置列宽为最大长度比例
                        worksheet.column_dimensions[col_name].width = value

                if os.path.exists('参考 各支部递交入党申请书人数') is False:
                        print("文件夹不存在")
                        os.mkdir("参考 各支部递交入党申请书人数")

                workbook.save('参考 各支部递交入党申请书人数/{} 递交入党申请书人数（2021上）.xlsx'.format(zhibu_list[zb]))


if __name__ == '__main__':
        test_file_book()
        test_file_count()
